from openpyxl import load_workbook,Workbook
from openpyxl.styles import PatternFill
wb=load_workbook('text1.xlsx')
sh=wb.active
index=[]#记录那个数据是重复
tmp_list=[]#记录没有重复的数据

for i,cell in enumerate(sh['B']):#enumerate函数用于将一个可遍历的数据对象(如列表、元组或字符串)组合为一个索引序列，同时列出数据和数据下标
    flag=cell.value not in tmp_list
    if flag:
        tmp_list.append(cell.value)
    else:
        index.append(i)

fill=PatternFill('solid','AEEEEE')
for i,row in enumerate(sh.rows):#遍历行
    if i in index:#i 是否为重复数据行
        for cell in row:
            cell.fill=fill
        print(f'第{i}行是重复数据')
wb.save('D:\\MoneyLoveLive\\excel快速查重\\查重后表.xlsx')
