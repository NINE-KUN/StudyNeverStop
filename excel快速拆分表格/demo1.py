from openpyxl import load_workbook
wb=load_workbook('D:\\MoneyLoveLive\\excel快速拆分表格\\test1.xlsx')
sh=wb.active
sh1=wb.create_sheet('单价高于1元')
sh2=wb.create_sheet('单价低于1元')
data1=[]
data2=[]
for row in sh.rows:
    num=row[3].value
    if num<=1:
        data1.append(row)
    else:
        data2.append(row)
for d in data1:
    tmp_lst=[]
    for tmp in d:
        tmp_lst.append(tmp.value)
    sh2.append(tmp_lst)
for d in data2:
    tmp_lst=[]
    for tmp in d:
        tmp_lst.append(tmp.value)
    sh1.append(tmp_lst)
wb.save('D:\\MoneyLoveLive\\excel快速拆分表格\\test1.xlsx')





