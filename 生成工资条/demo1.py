from openpyxl import load_workbook,Workbook
wb=load_workbook('工资数据.xlsx')
sh=wb.active
title=None
for i,row in enumerate(sh.rows):#enumerate() 函数用于将一个可遍历的数据对象(如列表、元组或字符串)组合为一个索引序列，同时列出数据和数据下标，一般用在 for 循环当中。
    tem_lst=[]
    for cell in row:
        tem_lst.append(cell.value)
    if i==0:#判断是不是title数据（第一行）
        title=tem_lst
    else:
        #创建excel文件
        tmp_wb=Workbook()
        tmp_sh=tmp_wb.active
        tmp_sh.append(title)
        tmp_sh.append(tem_lst)
        tmp_wb.save(f'D:\\MoneyLoveLive\\生成工资条\\{i}_{tem_lst[1]}.xlsx')