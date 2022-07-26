from openpyxl import Workbook,load_workbook
from datetime import date
def crear_data():
    wb=Workbook()
    sh=wb.active
    rows=[
        ['Data','姓名','打卡时间'],
        [date(2022, 3, 10),'钱坤','18:59'],
        [date(2022, 3, 11), '邹涛', '18:20'],
        [date(2022, 3, 12), '魏万卉', '18:14'],
        [date(2022, 3, 14), '钱坤', '18:34'],
        [date(2022, 3, 15), '邹涛', '19:56'],
        [date(2022, 3, 16), '刘嘉瑞', '18:59']
    ]
    for row in rows:
        sh.append(row)
    wb.save('D:\\MoneyLoveLive\\excel快速统计\\test1.xlsx')

def statistics():
    wb=load_workbook('test1.xlsx')
    sh=wb.active
    data=[]
    for i in range(2,sh.max_row+1):#从第二行开始遍历忽略表头 以最后一行结尾 因为max_row不包括最后一行所以+1
        tmp_lst=[]#默认为3列
        for j in range(1,sh.max_column+1):
            tmp_lst.append(sh.cell(i,j).value)

        #统计操作
        h,m=tmp_lst[2].split(':')#根据：通过split拆分时和分
        full=int(h)*60+int(m)#算出了上班总共用时
        result_date=full-18*60#总上班用时-正常下班用时
        tmp_lst.append(result_date)#将统计数据加入列表
        tmp_lst[0]=tmp_lst[0].date()#让第一行数据重新复位
        data.append(tmp_lst)
    new_wb=Workbook()
    new_sh=new_wb.active
    for row in data:
        new_sh.append(row)
    new_wb.save('D:\\MoneyLoveLive\\excel快速统计\\统计后的数据.xlsx')

if __name__ == '__main__':
    crear_data()
    statistics()