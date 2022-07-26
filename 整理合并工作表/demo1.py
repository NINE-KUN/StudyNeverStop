#操作excel
from openpyxl import load_workbook,Workbook
#打开excel
wb=load_workbook('D:\\MoneyLoveLive\\整理合并工作表\\test1.xlsx')
#激活多个sheet
sh1=wb['单价低于1元']
sh2=wb['单价高于1元']
all=[]#存储所有的数据
#读取数据
for row in sh1.rows:#获取所有行
    temp_lst=[]#专门用来存储一行的值
    for cell in row:#获取所有列
        temp_lst.append(cell.value)#获取单元格的值
    all.append(temp_lst)#整合所有数据
#合并几个就读取几个 第二个sheet
for row in sh2.rows:
    temp_lst=[]
    for cell in row:
        temp_lst.append(cell.value)
    all.append(temp_lst)#整合所有数据
#保存说有的数据
wb=Workbook()
sh=wb.active
for row in all:#遍历所以的数据
    sh.append(row)
wb.save('D:\\MoneyLoveLive\\整理合并工作表\\合并后文件.xlsx')
