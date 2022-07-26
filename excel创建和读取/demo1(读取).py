import openpyxl

def open():
    from openpyxl import load_workbook
    wb=load_workbook('D:\\MoneyLoveLive\\excel创建和读取\\test1.xlsx')

    sh1=wb.active#默认激活第一个工作表（sheet1 ）
    sh2=wb["Sheet1"]#选择激活工作表
    print(sh1==sh2)
open()
print('--------------------获取所有工作表名-----------------------')
def show_sheets():#如果不知道工作表名称获取所有工作表名
    from openpyxl import  load_workbook
    wb=load_workbook('D:\\MoneyLoveLive\\excel创建和读取\\test1.xlsx')
    print(wb.sheetnames)
    for sh in wb:
        print(sh.title)
show_sheets()
print('--------------------------------通过切片的方式获取表中的指定单元格数据--------------------------------------')
def get_one_value():
    from  openpyxl import  load_workbook #调用模块方法
    wb=load_workbook('D:\\MoneyLoveLive\\excel创建和读取\\test1.xlsx')#打开文件
    sh1=wb['Sheet1']#读取表名
    v1=sh1.cell(1,1)#获取表中的数据
    print(v1)#打印的cell值
    print(v1.value)
    v2=sh1['A2'].value#获取单个指定位置的value值
    print(v2)
    v3s=sh1['A1:B3']
    print(v3s)
    for item in v3s:#遍历得到每一行的单元格元组
        for cell in item:#遍历得到每一行的每一个单元格对象
            print(cell.value)#打印单元格对象的值
get_one_value()
print('-------------------------获取表中所有行或所有列的数据-------------------------------')
def get_all_value():#获取所有数据
    from  openpyxl import  load_workbook
    wb=load_workbook('D:\\MoneyLoveLive\\excel创建和读取\\test1.xlsx')
    sh=wb['Sheet1']
    for row in sh.rows:#遍历得到每一行的单元格元组
        for cell in row:#遍历得到每一行的每一个单元格对象
            print(cell.value)
    for columns in sh.columns:#通过遍历获取每一列的数据
        for cell in columns:
            print(cell.value)
get_all_value()