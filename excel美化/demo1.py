from openpyxl import Workbook
from openpyxl.styles import Font,colors,Alignment,PatternFill

wb=Workbook()
sh=wb.active
#font=Font(name='微软雅黑',size=30,bold=True,italic=True,color=colors.BLUE)#创建字体样式对象size大小 bold加粗 italic斜体 color
font=Font(name='微软雅黑',size=30,bold=True,italic=True,color='FFC1C1')#创建字体样式对象size大小 bold加粗 italic斜体 color颜色
sh['B2']='Hello'
sh['B2'].font=font

'''修改单元格'''
'C6'
#首先定位单元格
sh.row_dimensions[6].height=30#设置高度
sh.column_dimensions['C'].width=15#设置宽度
sh['C6']='python'
sh['C6'].alignment=Alignment(horizontal='right',vertical='top')#horizontal修改左右对齐vertical修改上下对齐

sh['h6'].fill=PatternFill('solid','CDCDC1')#solid填充满

wb.save('D:\\MoneyLoveLive\\excel美化\\text1.xlsx')