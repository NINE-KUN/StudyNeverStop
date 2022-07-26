#引包
from openpyxl import load_workbook
#打开文件
wb=load_workbook('text1.xlsx')
#激活sheet
sh1=wb.active
#读取数据
all=[]
for row in sh1.iter_rows(min_row=2,min_col=2):#从第二行第二列开始遍历
    tmp_lst=[]
    for cell in row:
        tmp_lst.append(cell.value)
    all.append(tmp_lst)
#汇总
all_count=[]#为了记录每一行数据和
for a in all:
    count=0 #单行数据和
    for num in a:
        tmp=str(num)#转换成str，是为了获取判断是否数字的问题
        if tmp.isdigit():#进行判断是否数字
            count +=num
    all_count.append(count)
    print(all_count)

print(sh1.max_row)#最高行
print(sh1.max_column)#最高列
colum_num=sh1.max_column+1#记录应该将数据写到那一列
sh1.cell(1,sh1.max_column+1).value='汇总'
row_num=2
for num in all_count:#获取所有的统计数据
    sh1.cell(row_num,colum_num).value=num
    row_num+=1
#保存
wb.save('D:\\MoneyLoveLive\\汇总excel\\统计后文件.xlsx')