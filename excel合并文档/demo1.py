#引入包
from openpyxl import load_workbook,Workbook
#打开excel文件
wb1=load_workbook('text1.py.xlsx')
wb2=load_workbook('text2.py.xlsx')
wb3=load_workbook('text3.py.xlsx')
#合并几个打开几个
#激活sheet
sh1=wb1.active
sh2=wb2.active
sh3=wb3.active
#定义变量存储所有数据
all=[]
#读取数据并保存
for row in sh1.rows:
    tmp_lst=[]#保存成一条完整的数据
    for cell in row:
        tmp_lst.append(cell.value)
    all.append(tmp_lst)
for row in sh2.rows:
    tmp_lst=[]#保存成一条完整的数据
    for cell in row:
        tmp_lst.append(cell.value)
    all.append(tmp_lst)
for row in sh3.rows:
    tmp_lst=[]#保存成一条完整的数据
    for cell in row:
        tmp_lst.append(cell.value)
    all.append(tmp_lst)
#保存数据
new_wb=Workbook()
new_sh=new_wb.active#new_wb.create_sheet('video)创建有名称的sheet

for row in all:#依次获取所有数据的每一条
    new_sh.append(row)
new_wb.save('D:\\MoneyLoveLive\\excel合并文档\\合并后excel.xlsx')

'''方法二'''
#打开excel文件
wb1=load_workbook('text1.py.xlsx')
wb2=load_workbook('text2.py.xlsx')
wb3=load_workbook('text3.py.xlsx')
#合并几个打开几个
#激活sheet
sh1=wb1.active
sh2=wb2.active
sh3=wb3.active
all_sheet=[sh1,sh2,sh3]
#定义变量存储所有数据
all=[]
#读取数据并保存
for sheet in all_sheet:
    for row in sheet.rows:
        tmp_lst=[]
        for cell in row:
            tmp_lst.append(cell.value)
        all.append(tmp_lst)

#保存数据
new_wb=Workbook()
new_sh=new_wb.active#new_wb.create_sheet('video)创建有名称的sheet

for row in all:#依次获取所有数据的每一条
    new_sh.append(row)
new_wb.save('D:\\MoneyLoveLive\\excel合并文档\\方法二合并后excel.xlsx')

